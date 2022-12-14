from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class XlsxWriter:
    """
        Class for writing data to YMarket's workbook
    """
    template: Workbook
    sheet: Worksheet
    category_name: str

    def __init__(self, data, category_name) -> None:
        self.read_template()
        self.category_name = category_name
        self.write_new_data(data)
        self.save()

    def read_template(self) -> None:
        """
          Reads YMarket template
        :return:
        """
        self.template = load_workbook('template.xlsx')
        worksheets: list[Worksheet] = self.template.worksheets
        self.sheet = worksheets[1]

    def write_new_data(self, data) -> None:
        """
          Writing data method
        :param data: list of offer Objects
        :return:
        """
        for index, offer in enumerate(data):
            self.sheet.cell(index+5, 3).value = offer.SKU
            self.sheet.cell(index+5, 4).value = offer.name
            self.sheet.cell(index+5, 5).value = offer.img
            self.sheet.cell(index+5, 6).value = offer.get_description()
            self.sheet.cell(index+5, 7).value = self.category_name
            self.sheet.cell(index+5, 8).value = offer.brand
            self.sheet.cell(index+5, 17).value = offer.get_detail_str()
            self.sheet.cell(index+5, 18).value = offer.home_url
            self.sheet.cell(index+5, 20).value = offer.price
            self.sheet.cell(index+5, 21).value = offer.get_non_sale_price()
            self.sheet.cell(index+5, 22).value = offer.currency
            self.sheet.cell(index+5, 29).value = 4
            self.sheet.cell(index+5, 30).value = 4
            self.sheet.cell(index+5, 31).value = offer.availability

    def save(self):
        self.template.save(self.category_name.replace(" ", "_")+".xlsx")
Ed$s%5g5l>mYj9


K0&0ox6#T%
  "
/usr/lib/jvm/java-1.11.0-openjdk-amd64

/root/YMARKET_ultradar


[Unit]
Description=Apache Tomcat Web Application Container
After=network.target

[Service]
Type=forking

Environment=JAVA_HOME=/usr/lib/jvm/java-1.11.0-openjdk-amd64
Environment=CATALINA_PID=/opt/tomcat/temp/tomcat.pid
Environment=CATALINA_HOME=/opt/tomcat
Environment=CATALINA_BASE=/opt/tomcat
Environment='CATALINA_OPTS=-Xms512M -Xmx1024M -server -XX:+UseParallelGC'
Environment='JAVA_OPTS=-Djava.awt.headless=true -Djava.security.egd=file:/dev/./urandom'

ExecStart=/opt/tomcat/bin/startup.sh
ExecStop=/opt/tomcat/bin/shutdown.sh

User=tomcat
Group=tomcat
UMask=0007
RestartSec=10
Restart=always

[Install]
WantedBy=multi-user.target