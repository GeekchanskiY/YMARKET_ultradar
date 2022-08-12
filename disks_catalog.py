import ftplib
from VARS import ftp_login, ftp_url, ftp_password
import xml.etree.ElementTree as ET
from openpyxl import Workbook


def getfile(ftp_client, filename):
    ftp_client.retrbinary("RETR " + filename, open(filename, 'wb').write)


def main():
    ftp = ftplib.FTP(ftp_url)
    ftp.login(ftp_login, ftp_password)

    getfile(ftp, 'Base.xml')

    ftp.quit()

    tree = ET.parse("Base.xml")
    root = tree.getroot()
    wb = Workbook()
    sheet = wb.create_sheet("Disks")
    sheet.cell(row=1, column=1).value = "code"
    sheet.cell(row=1, column=2).value = "prcode"
    sheet.cell(row=1, column=3).value = "articul"
    sheet.cell(row=1, column=4).value = "picture"
    sheet.cell(row=1, column=5).value = "usn"
    sheet.cell(row=1, column=6).value = "stock"
    sheet.cell(row=1, column=7).value = "brand"
    sheet.cell(row=1, column=8).value = "auto"
    sheet.cell(row=1, column=9).value = "model"
    sheet.cell(row=1, column=10).value = "color"
    sheet.cell(row=1, column=11).value = "width"
    sheet.cell(row=1, column=12).value = "diameter"
    sheet.cell(row=1, column=13).value = "bolts_count"
    sheet.cell(row=1, column=14).value = "bolts_spaicing"
    sheet.cell(row=1, column=15).value = "bolts_spaicing2"
    sheet.cell(row=1, column=16).value = "et"
    sheet.cell(row=1, column=17).value = "dia"
    sheet.cell(row=1, column=18).value = "price"
    sheet.cell(row=1, column=19).value = "count"
    sheet.cell(row=1, column=20).value = "countM"
    sheet.cell(row=1, column=11).value = "rrc"
    for index, el in enumerate(root.findall("gd")):
        row_num = 2 + index
        sheet.cell(row=row_num, column=1).value = el.find("code").text
        sheet.cell(row=row_num, column=2).value = el.find("PrCode").text
        sheet.cell(row=row_num, column=3).value = el.find("Articul").text
        sheet.cell(row=row_num, column=4).value = el.find("Picture").text
        sheet.cell(row=row_num, column=5).value = el.find("USN").text
        sheet.cell(row=row_num, column=6).value = el.find("Stock").text
        sheet.cell(row=row_num, column=7).value = el.find("brand").text
        sheet.cell(row=row_num, column=8).value = el.find("auto").text
        sheet.cell(row=row_num, column=9).value = el.find("model").text
        sheet.cell(row=row_num, column=10).value = el.find("color").text
        sheet.cell(row=row_num, column=11).value = el.find("width").text
        sheet.cell(row=row_num, column=12).value = el.find("diameter").text
        sheet.cell(row=row_num, column=13).value = el.find("bolts_count").text
        sheet.cell(row=row_num, column=14).value = el.find("bolts_spaicing").text
        sheet.cell(row=row_num, column=15).value = el.find("bolts_spaicing2").text
        sheet.cell(row=row_num, column=16).value = el.find("et").text
        sheet.cell(row=row_num, column=17).value = el.find("dia").text
        sheet.cell(row=row_num, column=18).value = el.find("price").text
        sheet.cell(row=row_num, column=19).value = el.find("count").text
        sheet.cell(row=row_num, column=20).value = el.find("countM").text
        sheet.cell(row=row_num, column=11).value = el.find("rrc").text
    wb.save("output.xlsx")


if __name__ == '__main__':
    main()
